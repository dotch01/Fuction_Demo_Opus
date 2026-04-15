"""
generate_scene.py — Excel → SceneConfig.json 轉換工具

場景物件互動（太空船等）的對話內容由此工具輸出。

───────────────────────────────────────────────────
Excel 格式（Sheet「場景物件」）：
  A  scene_id    場景識別碼       spaceship
  B  object_id   物件識別碼       console
  C  label       物件顯示名稱     控制台
  D  pos_x       X 位置(px)      -300   (相對螢幕中心，向右為正)
  E  pos_y       Y 位置(px)       0     (相對螢幕中心，向上為正)
  F  width       寬度(px)         120
  G  height      高度(px)         120
  H  color_r     顏色紅 (0-255)   77
  I  color_g     顏色綠 (0-255)   102
  J  color_b     顏色藍 (0-255)   153
  K  speaker     說話者名稱       系統
  L  portrait    立繪識別碼       (留空=無立繪)
  M  text        對話內容         控制台運作正常。

規則：
  - 同一 object_id 的連續行 = 同一物件的多句對話（按順序播放）
  - 物件欄位（label, pos, size, color）只需在第一行填寫，後續行可留空
  - 同一 scene_id 的行屬於同一場景

───────────────────────────────────────────────────
使用方式：
  pip install openpyxl
  python generate_scene.py scene.xlsx ../Assets/StreamingAssets/SceneConfig.json
"""

import json
import sys
from collections import OrderedDict

try:
    import openpyxl
except ImportError:
    print("請先安裝 openpyxl: pip install openpyxl")
    sys.exit(1)


def parse_float(val, default=0.0):
    try:
        return float(val)
    except (TypeError, ValueError):
        return default


def parse_int(val, default=0):
    try:
        return int(float(val))
    except (TypeError, ValueError):
        return default


def to_unity_color(val_0_255):
    """將 0-255 整數轉換為 Unity 使用的 0.0-1.0 浮點數，保留兩位小數。"""
    return round(parse_int(val_0_255, 128) / 255.0, 3)


def main():
    if len(sys.argv) < 3:
        print(f"用法: python {sys.argv[0]} <input.xlsx> <output.json>")
        sys.exit(1)

    xlsx_path = sys.argv[1]
    json_path = sys.argv[2]

    wb = openpyxl.load_workbook(xlsx_path, read_only=True)

    if "場景物件" not in wb.sheetnames:
        print("找不到 Sheet「場景物件」，請確認 Excel 格式正確。")
        sys.exit(1)

    ws = wb["場景物件"]

    # scenes[scene_id][object_id] = { ...meta..., dialogues: [...] }
    scenes = OrderedDict()
    # 記錄物件的 meta（只取第一次遇到）
    object_meta = {}

    for row in ws.iter_rows(min_row=2, values_only=True):
        scene_id = str(row[0] or "").strip()
        object_id = str(row[1] or "").strip()
        if not scene_id or not object_id:
            continue

        if scene_id not in scenes:
            scenes[scene_id] = OrderedDict()

        if object_id not in scenes[scene_id]:
            scenes[scene_id][object_id] = {
                "objectId": object_id,
                "label": str(row[2] or "").strip(),
                "posX": parse_float(row[3]),
                "posY": parse_float(row[4]),
                "width": parse_float(row[5], 100),
                "height": parse_float(row[6], 100),
                "colorR": to_unity_color(row[7]),
                "colorG": to_unity_color(row[8]),
                "colorB": to_unity_color(row[9]),
                "dialogues": [],
            }

        text = str(row[12] or "").strip()
        if text:
            scenes[scene_id][object_id]["dialogues"].append({
                "speaker": str(row[10] or "").strip(),
                "portrait": str(row[11] or "").strip(),
                "text": text,
            })

    wb.close()

    output = {
        "scenes": [
            {
                "id": scene_id,
                "objects": list(objects.values()),
            }
            for scene_id, objects in scenes.items()
        ]
    }

    with open(json_path, "w", encoding="utf-8") as f:
        json.dump(output, f, ensure_ascii=False, indent=2)

    total_objects = sum(len(objs) for objs in scenes.values())
    print(f"完成: {len(scenes)} 場景, {total_objects} 物件 → {json_path}")
    for scene_id, objects in scenes.items():
        print(f"  [{scene_id}]")
        for obj_id, obj in objects.items():
            lines = len(obj["dialogues"])
            print(f"    {obj_id} ({obj['label']}) — {lines} 句對話")


if __name__ == "__main__":
    main()
