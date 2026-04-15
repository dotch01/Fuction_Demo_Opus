"""
generate_items_from_excel.py — Excel → ItemConfig.json 轉換工具

Excel 須有一個 Sheet，名稱為「道具」。
第一行為標題列，會被跳過。

───────────────────────────────────────────────────
Sheet「道具」欄位：
  A  id              道具 ID             coin_01
  B  name            道具名稱            金幣
  C  category        分類                Currency / Consumable / Quest / Collection
  D  description     說明文字            閃閃發光的金幣...
  E  scoreValue      分數價值            10
  F  stackable       可否堆疊            TRUE / FALSE
  G  maxStack        最大堆疊數          99
  H  spriteName      圖形名稱            circle / diamond / square
  I  color_r         顏色 R (0~1)        1.0
  J  color_g         顏色 G (0~1)        0.85
  K  color_b         顏色 B (0~1)        0.2
  L  color_a         顏色 A (0~1)        1.0
  M  spriteScale     圖形縮放            0.35
  N  effect_type     效果類型            RestoreHP / RestoreMP / DamageHP / DamageMP (留空=無效果)
  O  effect_value    效果數值            20 (留空=0)

───────────────────────────────────────────────────
使用方式：
  pip install openpyxl
  python generate_items_from_excel.py items.xlsx ../Assets/StreamingAssets/ItemConfig.json

  若省略參數：
  python generate_items_from_excel.py
  → 讀取同目錄 items.xlsx，輸出 ../Assets/StreamingAssets/ItemConfig.json
"""

import json
import sys
import os

try:
    import openpyxl
except ImportError:
    print("請先安裝 openpyxl: pip install openpyxl")
    sys.exit(1)


def parse_bool(val):
    if isinstance(val, bool):
        return val
    return str(val).strip().upper() in ("TRUE", "1", "YES", "是")


def parse_float(val, default=0.0):
    try:
        return float(val)
    except (TypeError, ValueError):
        return default


def parse_int(val, default=0):
    try:
        return int(val)
    except (TypeError, ValueError):
        return default


def main():
    # 參數處理
    script_dir = os.path.dirname(os.path.abspath(__file__))
    xlsx_path = sys.argv[1] if len(sys.argv) > 1 else os.path.join(script_dir, "items.xlsx")
    output_path = sys.argv[2] if len(sys.argv) > 2 else os.path.join(
        script_dir, "..", "Assets", "StreamingAssets", "ItemConfig.json"
    )

    if not os.path.exists(xlsx_path):
        print(f"找不到 Excel 檔案: {xlsx_path}")
        print()
        print("請建立 Excel 檔案，Sheet 名稱為「道具」，欄位如下：")
        print("  A: id, B: name, C: category, D: description, E: scoreValue,")
        print("  F: stackable, G: maxStack, H: spriteName, I: color_r, J: color_g,")
        print("  K: color_b, L: color_a, M: spriteScale, N: effect_type, O: effect_value")
        sys.exit(1)

    wb = openpyxl.load_workbook(xlsx_path, data_only=True)

    # 嘗試找 Sheet
    sheet_name = None
    for name in ["道具", "Items", "items", "Sheet1"]:
        if name in wb.sheetnames:
            sheet_name = name
            break

    if sheet_name is None:
        print(f"找不到 Sheet。可用 Sheet: {wb.sheetnames}")
        print("請確認 Sheet 名稱為「道具」或「Items」")
        sys.exit(1)

    ws = wb[sheet_name]
    items = []

    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if not row or not row[0]:
            continue  # 跳過空行

        item_id = str(row[0]).strip()
        name = str(row[1] or "").strip()
        category = str(row[2] or "Currency").strip()
        description = str(row[3] or "").strip()
        score_value = parse_int(row[4])
        stackable = parse_bool(row[5]) if len(row) > 5 and row[5] is not None else True
        max_stack = parse_int(row[6], 99) if len(row) > 6 else 99
        sprite_name = str(row[7] or "circle").strip() if len(row) > 7 else "circle"
        color_r = parse_float(row[8], 1.0) if len(row) > 8 else 1.0
        color_g = parse_float(row[9], 1.0) if len(row) > 9 else 1.0
        color_b = parse_float(row[10], 1.0) if len(row) > 10 else 1.0
        color_a = parse_float(row[11], 1.0) if len(row) > 11 else 1.0
        sprite_scale = parse_float(row[12], 0.3) if len(row) > 12 else 0.3

        effect_type = str(row[13] or "").strip() if len(row) > 13 else ""
        effect_value = parse_int(row[14]) if len(row) > 14 else 0

        item = {
            "id": item_id,
            "name": name,
            "category": category,
            "description": description,
            "scoreValue": score_value,
            "stackable": stackable,
            "maxStack": max_stack,
            "spriteName": sprite_name,
            "color": {"r": color_r, "g": color_g, "b": color_b, "a": color_a},
            "spriteScale": sprite_scale,
            "effect": None,
        }

        if effect_type and effect_type.lower() != "none":
            item["effect"] = {"type": effect_type, "value": effect_value}

        items.append(item)
        print(f"  [{row_idx}] {item_id:20s} {name:12s} ({category})")

    # 輸出 JSON
    output_dir = os.path.dirname(output_path)
    if output_dir and not os.path.exists(output_dir):
        os.makedirs(output_dir, exist_ok=True)

    config = {"items": items}
    with open(output_path, "w", encoding="utf-8") as f:
        json.dump(config, f, ensure_ascii=False, indent=2)

    print(f"\n✓ 產生 {len(items)} 個道具 → {os.path.abspath(output_path)}")


if __name__ == "__main__":
    main()
