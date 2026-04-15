"""
generate_planets_from_excel.py — Excel → PlanetConfig.json 轉換工具

Excel 須有兩個 Sheet，名稱分別為「星域」和「行星」。
第一行為標題列，會被跳過。

───────────────────────────────────────────────────
Sheet「星域」欄位：
  A  name          星域名稱          GEVURAH
  B  quadrant      所在象限 (1-4)    1
  C  center_x      中心 X 座標       25
  D  center_y      中心 Y 座標       75
  E  radius        星域半徑          18

───────────────────────────────────────────────────
Sheet「行星」欄位：
  A  id              行星 ID           EMETH-100
  B  defaultName     預設名稱          Nono_01  (非任務行星留空)
  C  description     描述              此行星...  (非任務行星留空)
  D  pos_x           X 座標            34.89
  E  pos_y           Y 座標            65.96
  F  scale           大小比例          0.32
  G  radius          半徑 (km)         4118
  H  mass            質量              29182
  I  temperature     溫度 (°C)         92
  J  waterPercent    水分 (%)          69
  K  earthSimilarity 地球相似度 (%)    13.89
  L  isQuestTarget   是否任務目標      TRUE / FALSE
  M  filterOnly      是否需要濾鏡      TRUE / FALSE
  N  questOrder      任務順序          1  (非任務行星填 0 或留空)

───────────────────────────────────────────────────
使用方式：
  pip install openpyxl
  python generate_planets_from_excel.py planets.xlsx ../Assets/StreamingAssets/PlanetConfig.json
"""

import json
import sys

try:
    import openpyxl
except ImportError:
    print("請先安裝 openpyxl: pip install openpyxl")
    sys.exit(1)


def parse_bool(val):
    if isinstance(val, bool):
        return val
    return str(val).strip().upper() in ("TRUE", "1", "YES", "是")


def parse_float(val, default=0.0):
    try:
        return float(val)
    except (TypeError, ValueError):
        return default


def parse_int(val, default=0):
    try:
        return int(float(val))
    except (TypeError, ValueError):
        return default


def load_star_systems(ws):
    systems = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        name = str(row[0] or "").strip()
        if not name:
            continue
        systems.append({
            "name": name,
            "quadrant": parse_int(row[1]),
            "center": {
                "x": parse_float(row[2]),
                "y": parse_float(row[3]),
            },
            "radius": parse_float(row[4]),
        })
    return systems


def load_planets(ws):
    planets = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        pid = str(row[0] or "").strip()
        if not pid:
            continue
        planet = {
            "id": pid,
            "defaultName": str(row[1] or "").strip(),
            "description": str(row[2] or "").strip(),
            "position": {
                "x": parse_float(row[3]),
                "y": parse_float(row[4]),
            },
            "scale": parse_float(row[5], 0.3),
            "radius": parse_int(row[6]),
            "mass": parse_int(row[7]),
            "temperature": parse_int(row[8]),
            "waterPercent": parse_int(row[9]),
            "earthSimilarity": parse_float(row[10]),
            "isQuestTarget": parse_bool(row[11]),
            "filterOnly": parse_bool(row[12]),
            "questOrder": parse_int(row[13]),
        }
        planets.append(planet)
    return planets


def main():
    if len(sys.argv) < 3:
        print(f"用法: python {sys.argv[0]} <input.xlsx> <output.json>")
        sys.exit(1)

    xlsx_path = sys.argv[1]
    json_path = sys.argv[2]

    wb = openpyxl.load_workbook(xlsx_path, read_only=True)

    if "星域" not in wb.sheetnames:
        print("找不到 Sheet「星域」，請確認 Excel 格式正確。")
        sys.exit(1)
    if "行星" not in wb.sheetnames:
        print("找不到 Sheet「行星」，請確認 Excel 格式正確。")
        sys.exit(1)

    star_systems = load_star_systems(wb["星域"])
    planets = load_planets(wb["行星"])
    wb.close()

    config = {
        "starSystems": star_systems,
        "planets": planets,
    }

    with open(json_path, "w", encoding="utf-8") as f:
        json.dump(config, f, ensure_ascii=False, indent=2)

    print(f"完成: {len(star_systems)} 星域, {len(planets)} 行星 → {json_path}")

    # 驗證任務行星
    quest_planets = [p for p in planets if p["isQuestTarget"]]
    print(f"  其中任務行星: {len(quest_planets)} 顆")
    for p in sorted(quest_planets, key=lambda x: x["questOrder"]):
        print(f"    questOrder={p['questOrder']}  {p['id']}  {p['defaultName']}")


if __name__ == "__main__":
    main()
